"""将两个 hooks 精度日志解析为 Excel 对比报告。

命令行：
    python parse_hooks_to_excel.py --base-log <base.log> --target-log <target.log> --output <result.xlsx>
    python parse_hooks_to_excel.py --base-log <base.log> --target-log <target.log> --output <result.xlsx> --mapping <map1.json> <map2.json>

输出包含 3~4 个工作表：
1) base_parsed：base 日志解析结果
2) target_parsed：target 日志解析结果
3) comparison：同名匹配 + 映射匹配（需 --mapping）的完整对比
4) compare_map（可选）：仅映射匹配的对比，需提供 --mapping 参数

映射 JSON 格式示例：
    [
      {"base": "visual\\.blocks\\.(\\d+)\\.mlp\\.act_fn", "target": "visual.blocks.\\\\1.mlp.act"},
      {"base": "language_model\\.model\\.layers\\.(\\d+)", "target": "language_model.layers.\\\\1"}
    ]
    base 为正则表达式（匹配组件核心路径，自动处理 [forward]: 前缀和 inputs/outputs 后缀），
    target 为替换串（支持 \\1 \\2 反向引用捕获组）
"""

import argparse
import ast
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


HEADER_RE = re.compile(
    r"^'(?P<name>[^']+)'\|\s*"
    r"(?:l1_norm\s+(?P<l1_norm>[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)\s*\|\s*)?"
    r"(?P<hash>[^|]+?)\s*\|\s*"
    r"(?P<dtype>[^|]+?)\s*\|\s*"
    r"(?P<shape>torch\.Size\([^)]*\))\s*\|\s*"
    r"continue:\s*(?:True|False)\s*\|\s*"
    r"mean\s+(?P<mean>[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)\s*\|\s*"
    r"sum\s+(?P<sum>[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)\s*\|\s*"
    r"(?:max\s+(?P<max_val>[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)\s*\|\s*)?"
    r"(?:min\s+(?P<min_val>[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)\s*\|\s*)?"
    r"Size\s+(?P<size>\d+)\s*\|\s*"
    r"Memory size:\s*[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?\s*GB\s*\|\s*"
    r"isNan\s+(?P<is_nan>True|False)\s*$"
)


@dataclass
class HookEntry:
    name: str
    l1_norm: float | None
    hash_value: str
    dtype: str
    shape: str
    mean: float
    sum_value: float
    max_value: float | None
    min_value: float | None
    size: int
    is_nan: bool
    tensor_preview: list[Any]


def _parse_preview_line(line: str) -> list[Any]:
    stripped = line.strip()
    if not (stripped.startswith("[") and stripped.endswith("]")):
        return []

    try:
        value = ast.literal_eval(stripped)
    except (SyntaxError, ValueError):
        return []

    if isinstance(value, list):
        return value[:10]
    return []


def parse_hook_log(log_path: Path) -> list[HookEntry]:
    lines = log_path.read_text(encoding="utf-8").splitlines()
    entries: list[HookEntry] = []
    i = 0

    while i < len(lines):
        line = lines[i].strip()
        match = HEADER_RE.match(line)
        if not match:
            i += 1
            continue

        preview = []
        if i + 1 < len(lines):
            preview = _parse_preview_line(lines[i + 1])

        entries.append(
            HookEntry(
                name=match.group("name"),
                l1_norm=float(match.group("l1_norm")) if match.group("l1_norm") is not None else None,
                hash_value=match.group("hash").strip(),
                dtype=match.group("dtype").strip(),
                shape=match.group("shape").strip(),
                mean=float(match.group("mean")),
                sum_value=float(match.group("sum")),
                max_value=float(match.group("max_val")) if match.group("max_val") is not None else None,
                min_value=float(match.group("min_val")) if match.group("min_val") is not None else None,
                size=int(match.group("size")),
                is_nan=match.group("is_nan") == "True",
                tensor_preview=preview,
            )
        )
        i += 2

    return entries


def _diff_pct(base_val: float, target_val: float) -> float:
    denominator = max(abs(base_val), abs(target_val), 1e-10)
    return abs(base_val - target_val) / denominator * 100.0


def build_comparison_rows(
    base_entries: list[HookEntry],
    target_entries: list[HookEntry],
    mapping_rules: list[dict[str, str]] | None = None,
) -> list[list[Any]]:
    """双指针顺序遍历 base 和 target，先尝试同名匹配，再尝试映射匹配。"""
    rows: list[list[Any]] = []

    target_used: set[int] = set()
    target_name_to_indices: dict[str, list[int]] = {}
    for idx, entry in enumerate(target_entries):
        target_name_to_indices.setdefault(entry.name, []).append(idx)

    target_cursor: dict[str, int] = {name: 0 for name in target_name_to_indices}

    def _try_match(name: str, base_size: int | None = None) -> tuple[int, Any] | None:
        indices = target_name_to_indices.get(name)
        if indices is None:
            return None
        cursor = target_cursor[name]
        while cursor < len(indices) and indices[cursor] in target_used:
            cursor += 1
        target_cursor[name] = cursor
        if cursor >= len(indices):
            return None

        scan = cursor
        while scan < len(indices):
            if indices[scan] in target_used:
                scan += 1
                continue
            if base_size is not None and target_entries[indices[scan]].size != base_size:
                scan += 1
                continue
            break

        if scan >= len(indices):
            return None

        target_row = indices[scan]
        target_used.add(target_row)
        return target_row, target_entries[target_row]

    for base_row, base_item in enumerate(base_entries):
        matched = _try_match(base_item.name, base_item.size)

        match_type = "same_name"
        target_name = base_item.name
        if matched is None and mapping_rules:
            for resolved in _resolve_target_names(base_item.name, mapping_rules):
                matched = _try_match(resolved, base_item.size)
                if matched is not None:
                    match_type = "mapped"
                    target_name = resolved
                    break

        if matched is not None:
            target_row, target_item = matched
            l1_norm_diff_pct = None
            mean_diff_pct = None
            sum_diff_pct = None
            max_diff_pct = None
            min_diff_pct = None
            if base_item.l1_norm is not None and target_item.l1_norm is not None:
                l1_norm_diff_pct = _diff_pct(base_item.l1_norm, target_item.l1_norm)
            if base_item.mean is not None and target_item.mean is not None:
                mean_diff_pct = _diff_pct(base_item.mean, target_item.mean)
            if base_item.sum_value is not None and target_item.sum_value is not None:
                sum_diff_pct = _diff_pct(base_item.sum_value, target_item.sum_value)
            if base_item.max_value is not None and target_item.max_value is not None:
                max_diff_pct = _diff_pct(base_item.max_value, target_item.max_value)
            if base_item.min_value is not None and target_item.min_value is not None:
                min_diff_pct = _diff_pct(base_item.min_value, target_item.min_value)

            rows.append(
                [
                    base_item.name,
                    target_name,
                    match_type,
                    base_row + 1,
                    target_row + 1,
                    base_item.hash_value == target_item.hash_value,
                    base_item.shape == target_item.shape,
                    base_item.l1_norm,
                    target_item.l1_norm,
                    l1_norm_diff_pct,
                    base_item.mean,
                    target_item.mean,
                    mean_diff_pct,
                    base_item.sum_value,
                    target_item.sum_value,
                    sum_diff_pct,
                    base_item.max_value,
                    target_item.max_value,
                    max_diff_pct,
                    base_item.min_value,
                    target_item.min_value,
                    min_diff_pct,
                    base_item.hash_value,
                    target_item.hash_value,
                    base_item.shape,
                    target_item.shape,
                ]
            )
            continue

        rows.append(
            [
                base_item.name,
                base_item.name,
                "unmatched",
                base_row + 1,
                None,
                False,
                False,
                base_item.l1_norm,
                None,
                None,
                base_item.mean,
                None,
                None,
                base_item.sum_value,
                None,
                None,
                base_item.max_value,
                None,
                None,
                base_item.min_value,
                None,
                None,
                base_item.hash_value,
                "",
                base_item.shape,
                "",
            ]
        )

    for target_row, target_item in enumerate(target_entries):
        if target_row in target_used:
            continue
        rows.append(
            [
                target_item.name,
                target_item.name,
                "unmatched",
                None,
                target_row + 1,
                False,
                False,
                None,
                target_item.l1_norm,
                None,
                None,
                target_item.mean,
                None,
                None,
                target_item.sum_value,
                None,
                None,
                target_item.max_value,
                None,
                None,
                target_item.min_value,
                None,
                "",
                target_item.hash_value,
                "",
                target_item.shape,
            ]
        )

    return rows


def _write_parsed_sheet(workbook: Workbook, sheet_name: str, entries: list[HookEntry]) -> None:
    sheet = workbook.create_sheet(title=sheet_name)
    sheet.append(
        [
            "name",
            "hash",
            "dtype",
            "shape",
            "l1_norm",
            "mean",
            "sum",
            "max",
            "min",
            "size",
            "is_nan",
            "tensor_preview",
        ]
    )

    for entry in entries:
        sheet.append(
            [
                entry.name,
                entry.hash_value,
                entry.dtype,
                entry.shape,
                entry.l1_norm,
                entry.mean,
                entry.sum_value,
                entry.max_value,
                entry.min_value,
                entry.size,
                entry.is_nan,
                json.dumps(entry.tensor_preview, ensure_ascii=False),
            ]
        )


def _add_pct_color_scale(sheet: Any) -> None:
    pct_suffix = "_diff_pct"
    pct_cols = [
        i + 1 for i, cell in enumerate(sheet[1]) if (cell.value or "").endswith(pct_suffix)
    ]
    last_row = sheet.max_row
    if last_row < 2:
        return

    rule = ColorScaleRule(
        start_type="num",
        start_value=0,
        start_color="63BE7B",
        mid_type="num",
        mid_value=5,
        mid_color="FFEB84",
        end_type="num",
        end_value=20,
        end_color="F8696B",
    )

    for col in pct_cols:
        col_letter = get_column_letter(col)
        range_str = f"{col_letter}2:{col_letter}{last_row}"
        sheet.conditional_formatting.add(range_str, rule)


def _write_comparison_sheet(workbook: Workbook, rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title="comparison")
    sheet.append(
        [
            "base_name",
            "target_name",
            "match_type",
            "base_row",
            "target_row",
            "hash_match",
            "shape_match",
            "base_l1_norm",
            "target_l1_norm",
            "l1_norm_diff_pct",
            "base_mean",
            "target_mean",
            "mean_diff_pct",
            "base_sum",
            "target_sum",
            "sum_diff_pct",
            "base_max",
            "target_max",
            "max_diff_pct",
            "base_min",
            "target_min",
            "min_diff_pct",
            "base_hash",
            "target_hash",
            "base_shape",
            "target_shape",
        ]
    )

    for row in rows:
        sheet.append(row)

    _add_pct_color_scale(sheet)


def _load_mapping_rules(mapping_path: Path) -> list[dict[str, str]]:
    raw = json.loads(mapping_path.read_text(encoding="utf-8"))
    if not isinstance(raw, list):
        raise ValueError("mapping JSON must be a list of {base, target} objects")
    for i, rule in enumerate(raw):
        if "base" not in rule or "target" not in rule:
            raise ValueError(f"mapping rule #{i} missing 'base' or 'target' key: {rule}")
    return raw


_HOOK_NAME_RE = re.compile(r"^\[forward\]: (.+?) (inputs|outputs)$")


def _strip_hook_prefix_suffix(name: str) -> tuple[str, str] | None:
    m = _HOOK_NAME_RE.match(name)
    if m is None:
        return None
    return m.group(1), m.group(2)


def _resolve_target_names(base_name: str, mapping_rules: list[dict[str, str]]) -> list[str]:
    base_stripped = _strip_hook_prefix_suffix(base_name)
    if base_stripped is not None:
        core_name, io_tag = base_stripped
    else:
        core_name = base_name
        io_tag = ""

    results: list[str] = []
    for rule in mapping_rules:
        base_pat = re.compile(rule["base"])
        m = base_pat.fullmatch(core_name)
        if m is None:
            continue

        target_core = rule["target"]
        for gi in range(min(len(m.groups()), 9), 0, -1):
            target_core = target_core.replace(f"\\{gi}", m.group(gi) or "")

        results.append(f"[forward]: {target_core} {io_tag}" if io_tag else target_core)

    return results


def build_mapped_comparison_rows(
    base_entries: list[HookEntry],
    target_entries: list[HookEntry],
    mapping_rules: list[dict[str, str]],
) -> list[list[Any]]:
    target_used: set[int] = set()
    target_name_to_indices: dict[str, list[int]] = {}
    for idx, entry in enumerate(target_entries):
        target_name_to_indices.setdefault(entry.name, []).append(idx)

    target_cursor: dict[str, int] = {name: 0 for name in target_name_to_indices}

    rows: list[list[Any]] = []

    for base_row, base_item in enumerate(base_entries):
        matched_target = None
        for target_name in _resolve_target_names(base_item.name, mapping_rules):
            indices = target_name_to_indices.get(target_name)
            if indices is None:
                continue

            cursor = target_cursor[target_name]
            while cursor < len(indices) and indices[cursor] in target_used:
                cursor += 1
            target_cursor[target_name] = cursor
            if cursor >= len(indices):
                continue

            scan = cursor
            while scan < len(indices):
                if indices[scan] in target_used:
                    scan += 1
                    continue
                if base_item.size != target_entries[indices[scan]].size:
                    scan += 1
                    continue
                break

            if scan >= len(indices):
                continue

            target_row = indices[scan]
            target_used.add(target_row)
            matched_target = (target_name, target_row, target_entries[target_row])
            target_cursor[target_name] = cursor + 1
            break

        if matched_target is None:
            continue

        target_name, target_row, target_item = matched_target

        l1_norm_diff_pct = None
        mean_diff_pct = None
        sum_diff_pct = None
        max_diff_pct = None
        min_diff_pct = None
        if base_item.l1_norm is not None and target_item.l1_norm is not None:
            l1_norm_diff_pct = _diff_pct(base_item.l1_norm, target_item.l1_norm)
        if base_item.mean is not None and target_item.mean is not None:
            mean_diff_pct = _diff_pct(base_item.mean, target_item.mean)
        if base_item.sum_value is not None and target_item.sum_value is not None:
            sum_diff_pct = _diff_pct(base_item.sum_value, target_item.sum_value)
        if base_item.max_value is not None and target_item.max_value is not None:
            max_diff_pct = _diff_pct(base_item.max_value, target_item.max_value)
        if base_item.min_value is not None and target_item.min_value is not None:
            min_diff_pct = _diff_pct(base_item.min_value, target_item.min_value)

        rows.append(
            [
                base_item.name,
                target_name,
                base_row + 1,
                target_row + 1,
                base_item.hash_value == target_item.hash_value,
                base_item.shape == target_item.shape,
                base_item.l1_norm,
                target_item.l1_norm,
                l1_norm_diff_pct,
                base_item.mean,
                target_item.mean,
                mean_diff_pct,
                base_item.sum_value,
                target_item.sum_value,
                sum_diff_pct,
                base_item.max_value,
                target_item.max_value,
                max_diff_pct,
                base_item.min_value,
                target_item.min_value,
                min_diff_pct,
                base_item.hash_value,
                target_item.hash_value,
                base_item.shape,
                target_item.shape,
            ]
        )

    return rows


def _write_mapped_comparison_sheet(workbook: Workbook, rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title="compare_map")
    sheet.append(
        [
            "base_name",
            "target_name",
            "base_row",
            "target_row",
            "hash_match",
            "shape_match",
            "base_l1_norm",
            "target_l1_norm",
            "l1_norm_diff_pct",
            "base_mean",
            "target_mean",
            "mean_diff_pct",
            "base_sum",
            "target_sum",
            "sum_diff_pct",
            "base_max",
            "target_max",
            "max_diff_pct",
            "base_min",
            "target_min",
            "min_diff_pct",
            "base_hash",
            "target_hash",
            "base_shape",
            "target_shape",
        ]
    )

    for row in rows:
        sheet.append(row)

    _add_pct_color_scale(sheet)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="将 hooks 日志解析为 Excel 对比文件")
    parser.add_argument("--base-log", required=True, help="基准日志文件路径")
    parser.add_argument("--target-log", required=True, help="目标日志文件路径")
    parser.add_argument("--output", required=True, help="输出 Excel 文件路径")
    parser.add_argument("--mapping", default=None, nargs="+", help="映射规则 JSON 文件路径（支持多个，用于异名组件对比）")
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    base_log = Path(args.base_log).expanduser().resolve()
    target_log = Path(args.target_log).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    if not base_log.exists():
        raise FileNotFoundError(f"base log not found: {base_log}")
    if not target_log.exists():
        raise FileNotFoundError(f"target log not found: {target_log}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    base_entries = parse_hook_log(base_log)
    target_entries = parse_hook_log(target_log)

    mapping_rules = None
    if args.mapping is not None:
        mapping_rules = []
        for mapping_arg in args.mapping:
            mapping_path = Path(mapping_arg).expanduser().resolve()
            if not mapping_path.exists():
                raise FileNotFoundError(f"mapping file not found: {mapping_path}")
            mapping_rules.extend(_load_mapping_rules(mapping_path))

    comparison_rows = build_comparison_rows(base_entries, target_entries, mapping_rules)

    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is not None:
        workbook.remove(active_sheet)

    _write_parsed_sheet(workbook, "base_parsed", base_entries)
    _write_parsed_sheet(workbook, "target_parsed", target_entries)
    _write_comparison_sheet(workbook, comparison_rows)

    if mapping_rules is not None:
        mapped_rows = build_mapped_comparison_rows(base_entries, target_entries, mapping_rules)
        _write_mapped_comparison_sheet(workbook, mapped_rows)
        print(f"Mapped comparison: {len(mapped_rows)} pairs matched via {len(mapping_rules)} rules")

    workbook.save(output_path)
    print(f"Wrote Excel comparison file: {output_path}")
    print(f"Parsed entries - base: {len(base_entries)}, target: {len(target_entries)}")


if __name__ == "__main__":
    main()
